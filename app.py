# app.py
import streamlit as st
import pandas as pd
from pathlib import Path
from PIL import Image
import uuid
import io
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
import requests

GITHUB_REPO   = st.secrets.get("GITHUB_REPO", "nukedashiso/order-system")
GITHUB_BRANCH = st.secrets.get("GITHUB_BRANCH", "main")
# GITHUB_TOKEN  = st.secrets.get("GITHUB_TOKEN", None)  # 私有 repo 才需要

def _gh_headers():
    return {"Authorization": f"Bearer {GITHUB_TOKEN}"} if GITHUB_TOKEN else {}

def gh_read_csv(path_in_repo: str) -> pd.DataFrame:
    url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{path_in_repo}"
    r = requests.get(url, headers=_gh_headers(), timeout=15)
    r.raise_for_status()
    return pd.read_csv(io.BytesIO(r.content), dtype=str)

def gh_read_excel(path_in_repo: str, sheet_name=0) -> pd.DataFrame:
    url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{path_in_repo}"
    r = requests.get(url, headers=_gh_headers(), timeout=15)
    r.raise_for_status()
    return pd.read_excel(io.BytesIO(r.content), sheet_name=sheet_name, engine="openpyxl")

# 用 GitHub 版本覆寫原本的 load_*：
def load_orders():
    try:
        return gh_read_csv("data/orders.csv")
    except Exception:
        return pd.DataFrame(columns=["order_id","user_name","note","created_at","is_paid"])

def load_order_items():
    try:
        df = gh_read_csv("data/order_items.csv").fillna("")
        if "qty" in df.columns: df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0).astype(int)
        if "unit_price" in df.columns: df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce").fillna(0.0)
        return df
    except Exception:
        return pd.DataFrame(columns=["order_id","item_name","qty","unit_price"])

# ========= 基本設定 =========
st.set_page_config(page_title="月會下午茶線上點餐", page_icon="🍱", layout="wide")
TZ = ZoneInfo("Asia/Taipei")

# 截單（可用 "18:00" 或 "2025/10/14, 18:00" / "2025-10-14, 18:00"）
CUTOFF = "2025/10/14 12:30"

# Excel 寫入位置（會持續累積）
EXCEL_PATH = "./exports/orders.xlsx"
ORDERS_WS = "Orders"
SUMMARY_WS = "Summary"

# 路徑
BASE = Path(".")
DATA_DIR = BASE / "data"
IMG_DIR = BASE /  "menus"   # 放兩張菜單圖片
DATA_DIR.mkdir(parents=True, exist_ok=True)
IMG_DIR.mkdir(parents=True, exist_ok=True)
if EXCEL_PATH:
    Path(EXCEL_PATH).parent.mkdir(parents=True, exist_ok=True)

# 本地 CSV（長期累積）
ORDERS_CSV = DATA_DIR / "orders.csv"           # order_id, user_name, note, created_at, is_paid
ORDER_ITEMS_CSV = DATA_DIR / "order_items.csv" # order_id, item_name, qty, unit_price

# ========= CSV I/O =========
def init_csv(path: Path, columns: list):
    if not path.exists():
        pd.DataFrame(columns=columns).to_csv(path, index=False)

init_csv(ORDERS_CSV, ["order_id","user_name","note","created_at","is_paid"])
init_csv(ORDER_ITEMS_CSV, ["order_id","item_name","qty","unit_price"])

def load_orders():
    if ORDERS_CSV.exists():
        return pd.read_csv(ORDERS_CSV, dtype=str)
    return pd.DataFrame(columns=["order_id","user_name","note","created_at","is_paid"])

def load_order_items():
    if ORDER_ITEMS_CSV.exists():
        df = pd.read_csv(ORDER_ITEMS_CSV, dtype=str).fillna("")
        if "qty" in df.columns:
            df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0).astype(int)
        if "unit_price" in df.columns:
            df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce").fillna(0.0)
        return df
    return pd.DataFrame(columns=["order_id","item_name","qty","unit_price"])

def save_orders(df): df.to_csv(ORDERS_CSV, index=False)
def save_order_items(df): df.to_csv(ORDER_ITEMS_CSV, index=False)

# ========= 截單判定 =========
def cutoff_state(cutoff_str: str):
    """
    回傳 (passed, msg)
    passed: True=已截單；False=仍可下單
    支援 'HH:MM' 或 'YYYY/MM/DD, %H:%M' / 'YYYY-%m-%d, %H:%M' / 'YYYY/MM/DD %H:%M'
    """
    now = datetime.now(TZ)
    try:
        if any(ch in cutoff_str for ch in ["/", "-", "年", "月"]):
            for fmt in ["%Y/%m/%d, %H:%M", "%Y-%m-%d, %H:%M", "%Y/%m/%d %H:%M"]:
                try:
                    cutoff = datetime.strptime(cutoff_str.strip(), fmt).replace(tzinfo=TZ)
                    break
                except ValueError:
                    continue
            else:
                raise ValueError("cutoff_str format error")
        else:
            hh, mm = map(int, cutoff_str.strip().split(":"))
            cutoff = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
    except Exception:
        cutoff = now.replace(hour=18, minute=0, second=0, microsecond=0)

    if now >= cutoff:
        return True, f"已截單（{cutoff.strftime('%Y/%m/%d %H:%M')}）"

    left = cutoff - now
    d = left.days
    h = left.seconds // 3600
    m = (left.seconds % 3600) // 60
    return False, f"距離截單剩餘 {d} 天 {h} 小時 {m} 分（{cutoff.strftime('%Y/%m/%d %H:%M')}）"

# ========= Excel I/O =========
def excel_append_order(excel_path: str, worksheet: str, row_values: list):
    """
    逐筆附加：時間, 訂單ID, 姓名, 明細(字串), 總額, 備註, 已收款
    """
    if not excel_path:
        return True, "excel_path 未設定，略過寫入"
    p = Path(excel_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    headers = ["時間","訂單ID","姓名","明細","總額","備註","已收款"]

    if not p.exists():
        df = pd.DataFrame([row_values], columns=headers)
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=worksheet or "Orders", index=False)
        return True, "OK"

    try:
        wb = load_workbook(p)
        wsname = worksheet or "Orders"
        ws = wb[wsname] if wsname in wb.sheetnames else wb.create_sheet(wsname)
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            ws.append(headers)
        ws.append(row_values)
        wb.save(p)
        return True, "OK"
    except Exception as e:
        return False, str(e)

def excel_upsert_summary(excel_path: str, worksheet: str, df: pd.DataFrame):
    """
    覆蓋寫入 Summary：品項, 單價, 數量, 金額
    """
    if not excel_path:
        return True, "excel_path 未設定，略過寫入"
    p = Path(excel_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    try:
        mode = "a" if p.exists() else "w"
        with pd.ExcelWriter(
            p, engine="openpyxl", mode=mode,
            if_sheet_exists=("replace" if mode=="a" else None)
        ) as writer:
            out = df.rename(columns={
                "item_name":"品項","unit_price":"單價","total_qty":"數量","amount":"金額"
            })
            out.to_excel(writer, sheet_name=worksheet or "Summary", index=False)
        return True, "OK"
    except Exception as e:
        return False, str(e)

# ========= 側邊欄：上傳菜單圖 =========
st.sidebar.title("🍽️ 線上點餐")
with st.sidebar.expander("菜單圖片維護", expanded=False):
    files = st.file_uploader("上傳圖片（jpg/png/jpeg）", type=["jpg","jpeg","png"], accept_multiple_files=True)
    if files:
        for f in files:
            Image.open(f).save(IMG_DIR / f"{uuid.uuid4().hex}.png")
        st.success("圖片已上傳！重新整理即可看到。")

mode = st.sidebar.radio("模式 / Mode", ["點餐模式", "確認模式"])

# ========= 前台點餐 =========
if mode == "點餐模式":
    st.title("📋 線上點餐")
    passed, msg = cutoff_state(CUTOFF)
    st.info(msg)

    # 顯示兩張菜單（取前兩張）＋ 點擊放大預覽（相容所有版本）
    imgs = sorted([p for p in IMG_DIR.glob("*") if p.suffix.lower() in [".jpg", ".jpeg", ".png"]])
    show_imgs = imgs[:2]
    HAS_MODAL = hasattr(st, "modal")  # 新舊版相容

    st.subheader("菜單")
    if not show_imgs:
        st.warning("尚未上傳菜單圖片（側邊欄可上傳）。")
    else:
        cols = st.columns(2)
        for i, p in enumerate(show_imgs):
            # 縮圖
            with cols[i % 2]:
                st.image(str(p), use_container_width=True, caption=f"菜單 {i+1}")
                if st.button(f"🔍 放大查看（菜單 {i+1}）", key=f"zoom_{i}"):
                    st.session_state["zoom_target"] = str(p)

            # 若此圖被選為放大
            if st.session_state.get("zoom_target") == str(p):
                img = Image.open(p)
                if HAS_MODAL:
                    # ✅ 支援新版：彈窗預覽
                    with st.modal(f"放大預覽｜菜單 {i+1}", key=f"modal_{i}", max_width=1200):
                        st.image(img, use_container_width=True)
                        buf = io.BytesIO(); img.save(buf, format="PNG"); buf.seek(0)
                        st.download_button(
                            "⬇️ 下載原圖",
                            data=buf.getvalue(),
                            file_name=f"menu_{i+1}.png",
                            mime="image/png",
                            use_container_width=True
                        )
                        if st.button("關閉", key=f"close_{i}", use_container_width=True):
                            st.session_state.pop("zoom_target", None)
                else:
                    # ✅ 舊版相容：頁內預覽
                    st.markdown(f"### 放大預覽｜菜單 {i+1}")
                    st.image(img, use_container_width=True)
                    buf = io.BytesIO(); img.save(buf, format="PNG"); buf.seek(0)
                    st.download_button(
                        "⬇️ 下載原圖",
                        data=buf.getvalue(),
                        file_name=f"menu_{i+1}.png",
                        mime="image/png",
                        key=f"dl_{i}"
                    )
                    if st.button("關閉預覽", key=f"close_fb_{i}"):
                        st.session_state.pop("zoom_target", None)
        st.divider()

    # ====== 填寫餐點（加入「版本號」避免清空不生效） ======
    st.subheader("填寫餐點")
    session_key = "rows_single_page_store"

    # 初始化 rows 與 版本號
    if session_key not in st.session_state:
        st.session_state[session_key] = [
            {"item_name": "", "unit_price": 0.0, "qty": 0},
            {"item_name": "", "unit_price": 0.0, "qty": 0},
        ]
    vkey = f"{session_key}_ver"
    if vkey not in st.session_state:
        st.session_state[vkey] = 0  # 版本號(用來刷新所有輸入元件)

    def add_row():
        st.session_state[session_key].append({"item_name": "", "unit_price": 0.0, "qty": 0})

    def clear_rows():
        # 重設兩列 + 版本號+1（強制所有輸入元件換新key）
        st.session_state[session_key] = [
            {"item_name": "", "unit_price": 0.0, "qty": 0},
            {"item_name": "", "unit_price": 0.0, "qty": 0},
        ]
        st.session_state[vkey] += 1
        # 同時把姓名/備註也清掉（選擇性）
        st.session_state["name_single_store"] = ""
        st.session_state["note_single_store"] = ""
        st.rerun()

    c1, c2, _ = st.columns([1, 1, 6])
    c1.button("新增", on_click=add_row, disabled=passed, use_container_width=True)
    c2.button("清空", on_click=clear_rows, disabled=passed, use_container_width=True)

    total = 0
    ver = st.session_state[vkey]  # 目前版本號

    with st.form(f"order_form_single_page_store_v{ver}", clear_on_submit=False):
        rows = st.session_state[session_key]
        for i, r in enumerate(rows):
            key_suffix = f"{ver}_{i}"
            a, b, c, d = st.columns([4, 2, 2, 2])
            r["item_name"]  = a.text_input("品項名稱", r["item_name"], key=f"nm_{key_suffix}", disabled=passed)
            r["unit_price"] = b.number_input("單價", min_value=0.0, step=1.0, value=float(r["unit_price"]), key=f"pr_{key_suffix}", disabled=passed)
            r["qty"]        = c.number_input("數量", min_value=0, step=1, value=int(r["qty"]), key=f"qt_{key_suffix}", disabled=passed)
            d.write(f"小計：${int(r['unit_price'] * r['qty'])}")
            total += int(r["unit_price"] * r["qty"])

        st.markdown(f"### 總計：${total}")
        # 姓名/備註也帶版本，避免殘留
        name = st.text_input("姓名/暱稱", st.session_state.get("name_single_store",""), key=f"name_single_store_{ver}", disabled=passed)
        note = st.text_input("備註（例如不要香菜／飲品糖冰）", st.session_state.get("note_single_store",""), key=f"note_single_store_{ver}", disabled=passed)
        submitted = st.form_submit_button("送出訂單", type="primary", use_container_width=True, disabled=passed)

    # 回存姓名/備註到固定鍵（讓下次顯示預設值）
    if 'name' in locals():
        st.session_state["name_single_store"] = name
    if 'note' in locals():
        st.session_state["note_single_store"] = note

    if submitted:
        if not name.strip():
            st.error("請輸入姓名/暱稱"); st.stop()
        valid_rows = [r for r in st.session_state[session_key] if r["item_name"].strip() and r["qty"]>0]
        if not valid_rows:
            st.error("請至少填一列有效餐點"); st.stop()

        # 寫入本地 CSV（長期保存）
        orders_df = load_orders()
        items_df  = load_order_items()
        oid = uuid.uuid4().hex[:12]
        now = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")

        orders_df = pd.concat([orders_df, pd.DataFrame([{
            "order_id": oid, "user_name": name.strip(),
            "note": note.strip(), "created_at": now, "is_paid": "否"
        }])], ignore_index=True)

        item_rows = []
        for r in valid_rows:
            item_rows.append({
                "order_id": oid,
                "item_name": r["item_name"].strip(),
                "qty": int(r["qty"]),
                "unit_price": float(r["unit_price"])
            })
        items_df = pd.concat([items_df, pd.DataFrame(item_rows)], ignore_index=True)
        save_orders(orders_df); save_order_items(items_df)

        # 寫入 Excel（Orders 附加一列）
        if EXCEL_PATH:
            detail_str = "; ".join([f"{r['item_name']}x{int(r['qty'])}@{int(r['unit_price'])}" for r in item_rows])
            ok, info = excel_append_order(
                EXCEL_PATH, ORDERS_WS,
                [now, oid, name.strip(), detail_str, total, note.strip(), "否"]
            )
            if ok: st.caption(f"已寫入 Excel：{EXCEL_PATH}（{ORDERS_WS}）")
            else:  st.warning(f"訂單已保存，但寫入 Excel 失敗：{info}")

        st.success(f"訂單送出成功！編號：{oid}")
        st.balloons()

# ========= 管理者模式 =========
else:
    st.title("🔧 確認模式")

    orders = load_orders()
    items  = load_order_items()

    if orders.empty:
        st.info("尚無訂單")
        st.stop()

    # 彙總（品項 + 單價）
    agg = items.groupby(["item_name","unit_price"], as_index=False).agg(total_qty=("qty","sum"))
    agg["amount"] = (agg["total_qty"] * agg["unit_price"]).astype(int)

    st.subheader("品項彙總")
    st.dataframe(
        agg.rename(columns={"item_name":"品項","unit_price":"單價","total_qty":"數量","amount":"金額"}),
        use_container_width=True
    )
    st.markdown(f"**總金額：** ${int(agg['amount'].sum())}")

    # 訂單列表 + 切換收款
    order_total = items.groupby("order_id", as_index=False).apply(
        lambda df: pd.Series({"order_total": int((df["qty"]*df["unit_price"]).sum())})
    ).reset_index(drop=True)
    orders = orders.merge(order_total, on="order_id", how="left").fillna({"order_total":0})
    orders = orders.sort_values("created_at", ascending=False)

    st.subheader("訂單列表")
    for _, od in orders.iterrows():
        with st.container(border=True):
            st.markdown(f"**訂單 {od['order_id']}**｜{od['created_at']}｜{od['user_name']}｜金額 ${int(od['order_total'])}")
            if str(od["note"]).strip():
                st.caption(f"備註：{od['note']}")
            c1, c2 = st.columns([1,1])
            with c1:
                st.write(f"已收款：{od['is_paid']}")
            with c2:
                if st.button(f"切換收款（{od['order_id']}）", key=f"pay_{od['order_id']}"):
                    all_orders = load_orders()
                    mask = (all_orders["order_id"] == od["order_id"])
                    cur = all_orders.loc[mask, "is_paid"].iloc[0]
                    all_orders.loc[mask, "is_paid"] = "否" if cur=="是" else "是"
                    save_orders(all_orders); st.success("已更新收款狀態"); st.rerun()

    # 一鍵同步彙總 → Excel（覆蓋 Summary）
    st.divider()
    st.subheader("一鍵同步彙總 → Excel")
    st.caption("覆蓋寫入 Summary 工作表。")
    if st.button("同步彙總"):
        if not EXCEL_PATH:
            st.warning("未設定 EXCEL_PATH")
        else:
            ok, info = excel_upsert_summary(EXCEL_PATH, SUMMARY_WS, agg)
            if ok: st.success(f"彙總已寫入：{EXCEL_PATH}（{SUMMARY_WS}）")
            else:  st.warning(f"寫入 Excel 失敗：{info}")

    # 即時生成並下載（不依賴磁碟）
    st.divider()
    st.subheader("下載 Excel")
    st.caption("包含 Orders（逐單）與 Summary（彙總）兩張工作表。")

    # 組「明細」與「總額」欄位
    detail = items.groupby("order_id").apply(
        lambda d: "; ".join([f"{r['item_name']}x{int(r['qty'])}@{int(r['unit_price'])}" for _, r in d.iterrows()])
    ).reset_index(name="明細")
    total = items.groupby("order_id").apply(
        lambda d: int((d["qty"]*d["unit_price"]).sum())
    ).reset_index(name="總額")

    export_orders = (
        orders.merge(detail, on="order_id", how="left")
              .merge(total,  on="order_id", how="left")
              .loc[:, ["created_at","order_id","user_name","明細","總額","note","is_paid"]]
              .rename(columns={"created_at":"時間","order_id":"訂單ID","user_name":"姓名","note":"備註","is_paid":"已收款"})
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_orders.to_excel(writer, sheet_name=ORDERS_WS, index=False)
        agg_out = agg.rename(columns={"item_name":"品項","unit_price":"單價","total_qty":"數量","amount":"金額"})
        agg_out.to_excel(writer, sheet_name=SUMMARY_WS, index=False)
    buf.seek(0)

    st.download_button(
        "⬇️ 下載 Excel（即時產生）",
        data=buf.getvalue(),
        file_name=f"orders_{datetime.now(TZ):%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )




